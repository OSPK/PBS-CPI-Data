import os
import openpyxl
from openpyxl import load_workbook
from tqdm import tqdm
import datetime

path = os.path.dirname(os.path.abspath(__file__))
done_path = os.path.join(path, "final")
xfolder = os.path.join(path, "excels")

all_items = {1: "wheat", 2: "wheat_flour_bag", 3: "rice_basmati_broken", 4: "rice_irri-6-9", 5: "bread_plain_med_size_(340-400_gm)", 6: "beef", 7: "mutton", 8: "chicken_live_(farm)", 9: "milk_fresh", 10: "curd", 11: "milk_powdered_nido", 12: "egg_hen_(farm)", 13: "mustard_oil", 14: "cooking_oil_(tin)", 15: "veg._ghee_(tin)", 16: "veg.ghee_loose", 17: "bananas", 18: "masoor_pulse_washed", 19: "moong_pulse_washed", 20: "mash_pulse_washed", 21: "gram_pulse_washed", 22: "potatoes", 23: "onions", 24: "tomatoes", 25: "sugar", 26: "gur", 27: "salt_powdered_loose_(lahori)", 28: "red_chillies_powder_loose", 29: "garlic", 30: "tea_(yellow_lable_200_gm)", 31: "cooked_beef_plate", 32: "cooked_dal_plate", 33: "tea_prepared_(sada)", 34: "cigarettes_k-2_(20's)", 35: "long_cloth", 36: "shirting", 37: "lawn", 38: "georgette", 39: "sandal_gents_bata", 40: "chappal_spng_bata", 41: "sandal_ladies_bata", 42: "electric_charges", 43: "gas_charges_upto_100m3", 44: "kerosene", 45: "firewood", 46: "energy_savor_14_wats", 47: "washing_soap_(200-250_gm.)", 48: "match_box", 49: "petrol", 50: "diesel", 51: "l.p.g.(_11_kg_cylender.)", 52: "tele_local_call", 53: "bath_soap_lifebouy_(standard)"}

headers = ["Date", "Islamabad","Rawalpindi","Gujranwala","Sialkot","Lahore","Faisalabad","Sargodha","Multan","Bahawalpur","Karachi","Hyderabad","Sukker","Larkana","Peshawar","Bannu","Quetta","Khuzdar","Average"]

def create_item_file(file, serial_number):
    files = []
    # r=root, d=directories, f = files
    for r, d, f in os.walk(xfolder):
        for file in f:
            if '.xlsx' in file:
                files.append(os.path.join(r, file))



    twb = openpyxl.Workbook()
    tsheet = twb.active
    tsheet.append(headers)
    tr = 2
    tc = 2

    month_dict = {"january": "1","february": "2","march": "3","april": "4","may": "5","june": "6","july": "7","august": "8","september": "9","october": "10","november": "11","december": "12"}

    for file in files:
        wb = load_workbook(filename = file)

        ws = wb['Page 1']

        max_row = ws.max_row+1
        # max_column = ws.max_column
        max_column = 22

        year = file.split("/")[-1].split("_")[0]
        month = file.split("/")[-1].split("_")[1].split(".xlsx")[0]
        month = month_dict[month]
        date = datetime.datetime.strptime(year+"/"+month, '%Y/%m').date()
        tsheet.cell(row=tr, column=1).value = date
        tsheet.cell(row=tr, column=1).number_format = 'YYYY/MM'

        for i in range(1,max_row):

            if ws.cell(row=i, column=1).value == serial_number:

                for j in range(4, max_column):

                    #print(ws.cell(row=i, column=j).value)
                    newc = ws.cell(row=i, column=j).value
                    tsheet.cell(row=tr, column=tc).value = newc
                    tc += 1
        tr += 1
        tc = 2

    twb.save(tomatoes_file)

#SINGLE
# serial_number = 53
# tomatoes_file = os.path.join(done_path, all_items[serial_number]+".xlsx")
# create_item_file(tomatoes_file, serial_number)

#Batch
for item in tqdm(range(1,54)):
    tomatoes_file = os.path.join(done_path, all_items[item]+".xlsx")
    serial_number = item
    create_item_file(tomatoes_file, serial_number)
